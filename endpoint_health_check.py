"""
PermitGrab - Endpoint Health Check & Tracker Update
=====================================================
Hits every active endpoint, grabs the most recent permit date, and:
  1. Logs per-city freshness (how many days since the last permit).
  2. Flags stale endpoints (no data in the last N days).
  3. Updates the master city tracker spreadsheet with last-permit dates.

Run standalone:
    python endpoint_health_check.py [--days-stale 7] [--update-xlsx path/to/tracker.xlsx]

Or import and call:
    from endpoint_health_check import run_health_check
    results = run_health_check()
"""

import json
import os
import sys
import time
import argparse
import requests
from datetime import datetime, timedelta
from collections import OrderedDict

# ââ Imports from the PermitGrab codebase ââââââââââââââââââââââââââââââââââââââ
# Adjust PYTHONPATH so we can import sibling modules when run standalone.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from city_source_db import get_active_cities, get_city_config
from city_configs import CITY_REGISTRY

# ââ Constants âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
DEFAULT_STALE_DAYS = 7          # Flag if no permit newer than this many days
REQUEST_TIMEOUT    = 30         # seconds
RETRY_LIMIT        = 2          # per endpoint
PAUSE_BETWEEN      = 0.25       # seconds between requests (be polite)

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "PermitGrab-HealthCheck/1.0",
    "Accept": "application/json",
})


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# Per-platform "latest permit date" fetchers
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# Each returns an ISO date string (YYYY-MM-DD) or None.

def _latest_socrata(config):
    """Query a Socrata SODA endpoint for the single most-recent record."""
    endpoint   = config["endpoint"]
    date_field = config["date_field"]
    params = {
        "$limit": 1,
        "$order": f"{date_field} DESC",
        "$select": date_field,
    }
    # Add city_filter if present (county/state datasets)
    city_filter = config.get("city_filter")
    if city_filter:
        ff = city_filter["field"]
        fv = city_filter["value"]
        params["$where"] = f"upper({ff}) = upper('{fv}')"

    resp = SESSION.get(endpoint, params=params, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    rows = resp.json()
    if not rows:
        return None
    raw = rows[0].get(date_field, "")
    return _parse_date(raw, config.get("date_format", "iso"))


def _latest_arcgis(config):
    """Query an ArcGIS FeatureServer for the most-recent record."""
    endpoint   = config["endpoint"]
    date_field = config["date_field"]
    date_fmt   = config.get("date_format", "date")

    where = "1=1"
    city_filter = config.get("city_filter")
    if city_filter:
        ff = city_filter["field"]
        fv = city_filter["value"]
        where = f"upper({ff}) = upper('{fv}')"

    params = {
        "where": where,
        "outFields": date_field,
        "resultRecordCount": 1,
        "orderByFields": f"{date_field} DESC",
        "f": "json",
    }
    resp = SESSION.get(endpoint, params=params, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    data = resp.json()

    if "error" in data:
        raise Exception(f"ArcGIS error: {data['error'].get('message', 'unknown')}")

    features = data.get("features", [])
    if not features:
        return None
    raw = features[0]["attributes"].get(date_field)
    return _parse_date(raw, date_fmt)


def _latest_ckan(config):
    """Query a CKAN DataStore for the most-recent record."""
    endpoint   = config["endpoint"]
    dataset_id = config["dataset_id"]
    date_field = config.get("date_field", "")
    if not date_field:
        return None

    params = {
        "resource_id": dataset_id,
        "limit": 1,
        "sort": f"{date_field} desc",
    }
    resp = SESSION.get(endpoint, params=params, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    data = resp.json()
    records = data.get("result", {}).get("records", [])
    if not records:
        return None
    raw = records[0].get(date_field, "")
    return _parse_date(raw, "iso")


def _latest_carto(config):
    """Query a CARTO SQL API for the most-recent record."""
    endpoint   = config["endpoint"]
    table_name = config.get("table_name", config.get("dataset_id", ""))
    date_field = config.get("date_field", "")
    if not date_field or not table_name:
        return None

    sql = f"SELECT {date_field} FROM {table_name} ORDER BY {date_field} DESC LIMIT 1"
    params = {"q": sql, "format": "json"}
    resp = SESSION.get(endpoint, params=params, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    data = resp.json()
    if "error" in data:
        raise Exception(f"CARTO error: {data['error']}")
    rows = data.get("rows", [])
    if not rows:
        return None
    raw = rows[0].get(date_field, "")
    return _parse_date(raw, "iso")


PLATFORM_FETCHERS = {
    "socrata": _latest_socrata,
    "arcgis":  _latest_arcgis,
    "ckan":    _latest_ckan,
    "carto":   _latest_carto,
}


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# Date parsing helpers
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def _parse_date(raw, date_format="iso"):
    """Normalise whatever the API returns into YYYY-MM-DD."""
    if raw is None or raw == "":
        return None

    # Epoch milliseconds (ArcGIS convention)
    if date_format == "epoch" or (isinstance(raw, (int, float)) and raw > 1_000_000_000_000):
        try:
            return datetime.utcfromtimestamp(int(raw) / 1000).strftime("%Y-%m-%d")
        except (ValueError, OSError):
            return None

    # "none" format â value is epoch ms stored as number
    if date_format == "none" and isinstance(raw, (int, float)):
        try:
            return datetime.utcfromtimestamp(int(raw) / 1000).strftime("%Y-%m-%d")
        except (ValueError, OSError):
            return None

    # ISO-ish string: "2026-03-24T14:30:00.000" or "2026-03-24"
    s = str(raw).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s[:26], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    # Last-ditch: grab YYYY-MM-DD with regex
    import re
    m = re.search(r"(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    return None


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# Main health-check driver
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def run_health_check(stale_days=DEFAULT_STALE_DAYS, verbose=True):
    """
    Hit every active endpoint once and return a list of result dicts:
      {
        "key":            "chicago",
        "name":           "Chicago",
        "state":          "IL",
        "platform":       "socrata",
        "endpoint":       "https://...",
        "latest_date":    "2026-03-24"  or None,
        "days_since":     1             or None,
        "status":         "fresh" | "stale" | "empty" | "error",
        "error":          ""            or error message,
        "response_ms":    234,
        "checked_at":     "2026-03-25T08:31:00",
      }
    """
    active_cities = get_active_cities()
    if not active_cities:
        print("[WARN] No active cities returned by get_active_cities(). Falling back to CITY_REGISTRY.")
        active_cities = [
            {**v, "source_key": k}
            for k, v in CITY_REGISTRY.items()
            if v.get("active", False)
        ]

    today = datetime.utcnow().date()
    results = []

    if verbose:
        print(f"\n{'='*72}")
        print(f"  PermitGrab Endpoint Health Check â {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print(f"  Active endpoints: {len(active_cities)}   Stale threshold: {stale_days} days")
        print(f"{'='*72}\n")

    for i, city in enumerate(active_cities, 1):
        key      = city.get("source_key", city.get("key", "unknown"))
        name     = city.get("name", key)
        state    = city.get("state", "??")
        platform = city.get("platform", "socrata")
        endpoint = city.get("endpoint", "")

        if verbose:
            print(f"  [{i:>3}/{len(active_cities)}] {name}, {state} ({platform})...", end=" ", flush=True)

        fetcher = PLATFORM_FETCHERS.get(platform)
        if not fetcher:
            result = {
                "key": key, "name": name, "state": state, "platform": platform,
                "endpoint": endpoint, "latest_date": None, "days_since": None,
                "status": "error", "error": f"Unknown platform: {platform}",
                "response_ms": 0, "checked_at": datetime.utcnow().isoformat(),
            }
            results.append(result)
            if verbose:
                print(f"SKIP (unknown platform)")
            continue

        last_error = ""
        latest_date = None
        response_ms = 0

        for attempt in range(RETRY_LIMIT):
            try:
                t0 = time.time()
                latest_date = fetcher(city)
                response_ms = int((time.time() - t0) * 1000)
                last_error = ""
                break
            except requests.exceptions.HTTPError as e:
                sc = e.response.status_code if e.response else "??"
                last_error = f"HTTP {sc}"
                response_ms = int((time.time() - t0) * 1000)
                if sc in (403, 404):
                    break  # Don't retry auth / not-found
            except requests.exceptions.Timeout:
                last_error = "Timeout"
                response_ms = REQUEST_TIMEOUT * 1000
            except requests.exceptions.ConnectionError as e:
                last_error = f"Connection error: {str(e)[:80]}"
                response_ms = int((time.time() - t0) * 1000)
            except Exception as e:
                last_error = str(e)[:120]
                response_ms = int((time.time() - t0) * 1000)
                break  # Don't retry unknown errors

            if attempt < RETRY_LIMIT - 1:
                time.sleep(1)

        # Determine status
        if last_error:
            status = "error"
            days_since = None
        elif latest_date is None:
            status = "empty"
            days_since = None
        else:
            try:
                d = datetime.strptime(latest_date, "%Y-%m-%d").date()
                days_since = (today - d).days
                status = "stale" if days_since > stale_days else "fresh"
            except ValueError:
                status = "error"
                days_since = None
                last_error = f"Bad date parse: {latest_date}"

        result = {
            "key": key, "name": name, "state": state, "platform": platform,
            "endpoint": endpoint, "latest_date": latest_date,
            "days_since": days_since, "status": status, "error": last_error,
            "response_ms": response_ms,
            "checked_at": datetime.utcnow().isoformat(),
        }
        results.append(result)

        if verbose:
            if status == "fresh":
                print(f"OK  latest={latest_date}  ({days_since}d ago)  {response_ms}ms")
            elif status == "stale":
                print(f"STALE  latest={latest_date}  ({days_since}d ago!)  {response_ms}ms")
            elif status == "empty":
                print(f"EMPTY  (no records returned)  {response_ms}ms")
            else:
                print(f"ERROR  {last_error}  {response_ms}ms")

        time.sleep(PAUSE_BETWEEN)

    # ââ Summary âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    if verbose:
        fresh = sum(1 for r in results if r["status"] == "fresh")
        stale = sum(1 for r in results if r["status"] == "stale")
        empty = sum(1 for r in results if r["status"] == "empty")
        error = sum(1 for r in results if r["status"] == "error")
        print(f"\n{'='*72}")
        print(f"  Summary:  {fresh} fresh  |  {stale} stale  |  {empty} empty  |  {error} error")
        print(f"{'='*72}")

        if stale + error > 0:
            print(f"\n  â   Problem endpoints:")
            for r in results:
                if r["status"] in ("stale", "error"):
                    tag = "STALE" if r["status"] == "stale" else "ERROR"
                    detail = f"last={r['latest_date']} ({r['days_since']}d)" if r["latest_date"] else r["error"]
                    print(f"     [{tag}] {r['name']}, {r['state']} â {detail}")

    return results


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# XLSX Tracker Updater
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def update_tracker_xlsx(results, xlsx_path):
    """
    Open the master city tracker spreadsheet and fill in:
      - Column H ("Last Permit Date") for every active city we checked
      - Optionally flag stale/error cities in the Notes column
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[ERROR] openpyxl not installed â cannot update XLSX")
        return False

    if not os.path.exists(xlsx_path):
        print(f"[ERROR] XLSX not found: {xlsx_path}")
        return False

    wb = load_workbook(xlsx_path)
    ws = wb["City Tracker"]

    # Build lookup from results: key -> result
    result_lookup = {r["key"]: r for r in results}
    # Also build by name+state for fuzzy matching
    name_lookup = {}
    for r in results:
        nk = f"{r['name'].lower()}|{r['state']}"
        name_lookup[nk] = r

    updated = 0
    for row in range(2, ws.max_row + 1):
        city_name  = ws.cell(row=row, column=2).value or ""
        state      = ws.cell(row=row, column=3).value or ""
        config_key = ws.cell(row=row, column=7).value or ""

        # Try to find a matching health-check result
        result = None
        if config_key and config_key in result_lookup:
            result = result_lookup[config_key]
        else:
            nk = f"{city_name.lower()}|{state}"
            result = name_lookup.get(nk)

        if result:
            # Update Last Permit Date (column H)
            if result["latest_date"]:
                ws.cell(row=row, column=8, value=result["latest_date"])
                updated += 1

            # If stale or error, append to notes
            if result["status"] in ("stale", "error"):
                current_notes = ws.cell(row=row, column=9).value or ""
                tag = f"[HEALTH {datetime.now().strftime('%m/%d')}] "
                if result["status"] == "stale":
                    tag += f"STALE â last permit {result['latest_date']} ({result['days_since']}d ago)"
                else:
                    tag += f"ERROR â {result['error']}"
                # Don't duplicate notes
                if tag not in current_notes:
                    ws.cell(row=row, column=9, value=tag if not current_notes else f"{current_notes}; {tag}")

    wb.save(xlsx_path)
    print(f"\n[XLSX] Updated {updated} 'Last Permit Date' cells in {xlsx_path}")
    return True


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# JSON report writer
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def save_json_report(results, path="data/health_check_report.json"):
    """Write full results to JSON for dashboarding / history."""
    report = {
        "checked_at": datetime.utcnow().isoformat(),
        "total_checked": len(results),
        "summary": {
            "fresh": sum(1 for r in results if r["status"] == "fresh"),
            "stale": sum(1 for r in results if r["status"] == "stale"),
            "empty": sum(1 for r in results if r["status"] == "empty"),
            "error": sum(1 for r in results if r["status"] == "error"),
        },
        "results": results,
    }
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w") as f:
        json.dump(report, f, indent=2)
    print(f"[JSON] Report saved to {path}")


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# CLI entry point
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def main():
    parser = argparse.ArgumentParser(description="PermitGrab Endpoint Health Check")
    parser.add_argument("--days-stale", type=int, default=DEFAULT_STALE_DAYS,
                        help=f"Days without data before flagging stale (default: {DEFAULT_STALE_DAYS})")
    parser.add_argument("--update-xlsx", type=str, default=None,
                        help="Path to master tracker XLSX to update with last permit dates")
    parser.add_argument("--json-report", type=str, default="data/health_check_report.json",
                        help="Path to save JSON health report (default: data/health_check_report.json)")
    parser.add_argument("--quiet", action="store_true",
                        help="Suppress per-endpoint output")
    args = parser.parse_args()

    results = run_health_check(stale_days=args.days_stale, verbose=not args.quiet)

    # Save JSON report
    save_json_report(results, args.json_report)

    # Update XLSX if requested
    if args.update_xlsx:
        update_tracker_xlsx(results, args.update_xlsx)

    # Exit code: non-zero if any errors or stale
    problems = sum(1 for r in results if r["status"] in ("stale", "error"))
    return 1 if problems > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
